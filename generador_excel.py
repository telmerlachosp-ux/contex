from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import date


def generar_excel_compra(datos_extraidos, resultado):
    """
    Recibe los datos identificados por la IA y el resultado del asiento,
    y devuelve un archivo Excel en memoria (listo para descargar),
    con formato de libro diario: el N° correlativo, la fecha y la glosa
    aparecen una sola vez por cada asiento.
    """
    wb = Workbook()

    # -------------------------
    # HOJA 1: RESUMEN
    # -------------------------
    hoja_resumen = wb.active
    hoja_resumen.title = "Resumen"

    hoja_resumen["A1"] = "RESUMEN DEL EJERCICIO"
    hoja_resumen["A1"].font = Font(bold=True, size=14)

    hoja_resumen["A3"] = "Base imponible"
    hoja_resumen["B3"] = datos_extraidos["base_imponible"]

    hoja_resumen["A4"] = "IGV"
    hoja_resumen["B4"] = datos_extraidos["igv"]

    hoja_resumen["A5"] = "Total"
    hoja_resumen["B5"] = datos_extraidos["total"]

    hoja_resumen["A6"] = "Condición de pago"
    hoja_resumen["B6"] = datos_extraidos["condicion_pago"]

    hoja_resumen.column_dimensions["A"].width = 20
    hoja_resumen.column_dimensions["B"].width = 15

    # -------------------------
    # HOJA 2: LIBRO DIARIO (ASIENTOS)
    # -------------------------
    hoja_asiento = wb.create_sheet("Asientos")

    encabezados = ["N° Corr", "Fecha", "Glosa", "Código", "Denominación", "Debe", "Haber"]
    hoja_asiento.append(encabezados)

    for celda in hoja_asiento[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(
            start_color="305496", end_color="305496", fill_type="solid"
        )
        celda.alignment = Alignment(horizontal="center", vertical="center")

    fecha_hoy = date.today().strftime("%d/%m/%Y")

    fila_actual = 2
    asiento_anterior = None
    fila_inicio_bloque = 2

    for cuenta in resultado["cuentas"]:
        numero_asiento = cuenta["asiento"]

        if numero_asiento != asiento_anterior:
            if asiento_anterior is not None:
                fila_fin_bloque = fila_actual - 1
                if fila_fin_bloque > fila_inicio_bloque:
                    for columna in [1, 2, 3]:
                        hoja_asiento.merge_cells(
                            start_row=fila_inicio_bloque,
                            start_column=columna,
                            end_row=fila_fin_bloque,
                            end_column=columna
                        )

            hoja_asiento.cell(row=fila_actual, column=1, value=numero_asiento)
            hoja_asiento.cell(row=fila_actual, column=2, value=fecha_hoy)
            hoja_asiento.cell(row=fila_actual, column=3, value=cuenta["glosa"])

            fila_inicio_bloque = fila_actual
            asiento_anterior = numero_asiento

        hoja_asiento.cell(row=fila_actual, column=4, value=cuenta["codigo"])
        hoja_asiento.cell(row=fila_actual, column=5, value=cuenta["cuenta"])
        hoja_asiento.cell(row=fila_actual, column=6, value=cuenta["debe"])
        hoja_asiento.cell(row=fila_actual, column=7, value=cuenta["haber"])

        fila_actual += 1

    fila_fin_bloque = fila_actual - 1
    if fila_fin_bloque > fila_inicio_bloque:
        for columna in [1, 2, 3]:
            hoja_asiento.merge_cells(
                start_row=fila_inicio_bloque,
                start_column=columna,
                end_row=fila_fin_bloque,
                end_column=columna
            )

    for fila in hoja_asiento.iter_rows(min_row=2, max_row=fila_actual - 1):
        for celda in fila:
            celda.alignment = Alignment(vertical="top", wrap_text=True)

    fila_total = fila_actual + 1
    hoja_asiento.cell(row=fila_total, column=5, value="TOTALES").font = Font(bold=True)
    hoja_asiento.cell(row=fila_total, column=6, value=resultado["debe"]).font = Font(bold=True)
    hoja_asiento.cell(row=fila_total, column=7, value=resultado["haber"]).font = Font(bold=True)

    fila_validacion = fila_total + 1
    texto_validacion = "CUADRADO" if resultado["cuadrado"] else "NO CUADRADO"
    hoja_asiento.cell(row=fila_validacion, column=5, value="Validación")
    hoja_asiento.cell(row=fila_validacion, column=6, value=texto_validacion)

    hoja_asiento.column_dimensions["A"].width = 10
    hoja_asiento.column_dimensions["B"].width = 12
    hoja_asiento.column_dimensions["C"].width = 45
    hoja_asiento.column_dimensions["D"].width = 12
    hoja_asiento.column_dimensions["E"].width = 28
    hoja_asiento.column_dimensions["F"].width = 15
    hoja_asiento.column_dimensions["G"].width = 15

    # -------------------------
    # GUARDAR EN MEMORIA
    # -------------------------
    archivo_en_memoria = BytesIO()
    wb.save(archivo_en_memoria)
    archivo_en_memoria.seek(0)

    return archivo_en_memoria


def generar_excel_multiples_asientos(datos_generales, asientos):
    """
    Genera un Excel a partir de una LISTA de asientos ya armados
    (cada uno con "tipo_asiento", "cuentas" y "validacion"), como
    los que devuelve generador_prestamos.py. A diferencia de
    generar_excel_compra, aquí cada asiento trae su propia
    validación y no usa el campo "asiento"/"glosa" por cuenta,
    sino "tipo_asiento" como título del bloque.
    """
    wb = Workbook()

    # -------------------------
    # HOJA 1: RESUMEN
    # -------------------------
    hoja_resumen = wb.active
    hoja_resumen.title = "Resumen"

    hoja_resumen["A1"] = "RESUMEN DEL EJERCICIO"
    hoja_resumen["A1"].font = Font(bold=True, size=14)

    fila_resumen = 3
    for clave, valor in datos_generales.items():
        hoja_resumen.cell(row=fila_resumen, column=1, value=str(clave))
        hoja_resumen.cell(row=fila_resumen, column=2, value=valor)
        fila_resumen += 1

    hoja_resumen.column_dimensions["A"].width = 22
    hoja_resumen.column_dimensions["B"].width = 20

    # -------------------------
    # HOJA 2: LIBRO DIARIO (ASIENTOS)
    # -------------------------
    hoja_asiento = wb.create_sheet("Asientos")

    encabezados = ["N° Corr", "Fecha", "Glosa", "Código", "Denominación", "Debe", "Haber"]
    hoja_asiento.append(encabezados)

    for celda in hoja_asiento[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(
            start_color="305496", end_color="305496", fill_type="solid"
        )
        celda.alignment = Alignment(horizontal="center", vertical="center")

    fecha_hoy = date.today().strftime("%d/%m/%Y")

    fila_actual = 2
    total_debe_general = 0
    total_haber_general = 0

    for indice, asiento in enumerate(asientos, start=1):
        glosa = asiento.get("tipo_asiento", "")
        cuentas = asiento.get("cuentas", [])
        fila_inicio_bloque = fila_actual

        for cuenta in cuentas:
            if fila_actual == fila_inicio_bloque:
                hoja_asiento.cell(row=fila_actual, column=1, value=indice)
                hoja_asiento.cell(row=fila_actual, column=2, value=fecha_hoy)
                hoja_asiento.cell(row=fila_actual, column=3, value=glosa)

            hoja_asiento.cell(row=fila_actual, column=4, value=cuenta["codigo"])
            hoja_asiento.cell(row=fila_actual, column=5, value=cuenta["cuenta"])
            hoja_asiento.cell(row=fila_actual, column=6, value=cuenta["debe"])
            hoja_asiento.cell(row=fila_actual, column=7, value=cuenta["haber"])

            total_debe_general += cuenta["debe"]
            total_haber_general += cuenta["haber"]

            fila_actual += 1

        fila_fin_bloque = fila_actual - 1
        if fila_fin_bloque > fila_inicio_bloque:
            for columna in [1, 2, 3]:
                hoja_asiento.merge_cells(
                    start_row=fila_inicio_bloque,
                    start_column=columna,
                    end_row=fila_fin_bloque,
                    end_column=columna
                )

    for fila in hoja_asiento.iter_rows(min_row=2, max_row=fila_actual - 1):
        for celda in fila:
            celda.alignment = Alignment(vertical="top", wrap_text=True)

    fila_total = fila_actual + 1
    hoja_asiento.cell(row=fila_total, column=5, value="TOTALES").font = Font(bold=True)
    hoja_asiento.cell(row=fila_total, column=6, value=round(total_debe_general, 2)).font = Font(bold=True)
    hoja_asiento.cell(row=fila_total, column=7, value=round(total_haber_general, 2)).font = Font(bold=True)

    diferencia_general = round(total_debe_general - total_haber_general, 2)
    fila_validacion = fila_total + 1
    texto_validacion = "CUADRADO" if diferencia_general == 0 else "NO CUADRADO"
    hoja_asiento.cell(row=fila_validacion, column=5, value="Validación")
    hoja_asiento.cell(row=fila_validacion, column=6, value=texto_validacion)

    hoja_asiento.column_dimensions["A"].width = 10
    hoja_asiento.column_dimensions["B"].width = 12
    hoja_asiento.column_dimensions["C"].width = 45
    hoja_asiento.column_dimensions["D"].width = 12
    hoja_asiento.column_dimensions["E"].width = 35
    hoja_asiento.column_dimensions["F"].width = 15
    hoja_asiento.column_dimensions["G"].width = 15

    # -------------------------
    # GUARDAR EN MEMORIA
    # -------------------------
    archivo_en_memoria = BytesIO()
    wb.save(archivo_en_memoria)
    archivo_en_memoria.seek(0)

    return archivo_en_memoria
